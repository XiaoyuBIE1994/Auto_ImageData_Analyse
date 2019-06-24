# -*- coding: utf-8 -*-
"""
Created on Tues Jun 13 14:07:05 2019

@author: xue
"""


import os
import xlrd
from openpyxl import load_workbook
import pandas as pd


def import_peak(workpath):
    peak_path = os.path.join(workpath, 'Peak_value.xlsx')
    wb = load_workbook(peak_path)
    sh =wb.worksheets[0]
    row_count = sh.max_row
    
    peak = {}
    wb = xlrd.open_workbook(peak_path)
    sh = wb.sheet_by_index(0)   
    for i in range(row_count):
        ID = sh.cell(i,0).value
        AppliedT = sh.cell(i,1).value
        peak[ID] = AppliedT
    return peak

def export_result(filename, workpath, result):
    
    peak_path = os.path.join(workpath, 'Peak_value.xlsx')
#     filename = 'SB_LGF-20&Plate_LGF20_Ep2_7_T1_1_001.PNG'
    data = pd.read_excel(peak_path)
    df = pd.DataFrame(data, columns =['ID'])
    df_index = df[df['ID']== filename[:-4]].index[0]
#     img_params = result_dir[filename[:-4]]

    wb = load_workbook(peak_path)
    sh =wb.worksheets[0]
    # print (img_params['x1'])
    sh.cell(df_index+2, 3).value = result['x0']
    sh.cell(df_index+2, 4).value = result['x1']
    sh.cell(df_index+2, 5).value = result['y0']
    sh.cell(df_index+2, 6).value = result['y1']
    sh.cell(df_index+2, 7).value = result['k1']
    sh.cell(df_index+2, 8).value = result['b1']
    sh.cell(df_index+2, 9).value = result['k2']
    sh.cell(df_index+2, 10).value = result['b2']
    sh.cell(df_index+2, 11).value = result['pt1_x'][0][0]
    sh.cell(df_index+2, 12).value = result['pt1_y'][0][0]
    sh.cell(df_index+2, 13).value = result['pt2_x'][0][0]
    # sh.cell(df_index+2, 14).value = img_params['pt2_y'][0][0]
    # print (df_index)
    wb.save(peak_path)
    wb.close()
    return




